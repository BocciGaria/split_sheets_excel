from openpyxl import *
from os import path, getcwd


def main():
    fname = input("ファイル：")
    abs_path = path.abspath(fname)
    fname_splited = path.splitext(path.basename(abs_path))

    if not path.exists(abs_path):
        print("指定されたファイルが見つかりません。")
        return

    wb_origin: Workbook = load_workbook(abs_path, read_only=True, keep_vba=True, keep_links=True)

    for ws_name in wb_origin.sheetnames:
        wb: Workbook = load_workbook(abs_path)
        for ws in wb.worksheets:
            if ws.title != ws_name:
                wb.remove(ws)
        path_to_export = path.join(getcwd(), fname_splited[0] + "_" + ws_name + fname_splited[1])
        wb.save(path_to_export)
        print("ファイル出力 --> '%s'" % path_to_export)


if __name__ == "__main__":
    main()
